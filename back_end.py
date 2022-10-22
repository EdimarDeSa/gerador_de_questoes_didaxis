from os.path import exists
import openpyxl as op
import selecao_de_diretorio


class FuncoesBackEnd:
    def inicia_excel(self):
        # Inicia a verificação de existência da planilha na pasta
        if exists(self.diretorio):
            # Caso exista a planilha, abre a planilha
            self.wb = op.load_workbook(self.diretorio)
            # Abre a primeira worksheet
            self.ws = self.wb[self.wb.sheetnames[0]]
        else:
            # Caso não exista a planilha, cria uma planilha
            self.wb = op.Workbook()
            # Abre a primeira worksheet
            self.ws = self.wb[self.wb.sheetnames[0]]
            # Salva o cabeçalho padrão que todos os bancos de questões devem ter
            self.ws.append(self.cabecalho)

    @property
    def cabecalho(self):
        # Cabeçalho padrão dos bancos de questões
        return [
            'ID',
            'TIPO',
            'PESO',
            'TEMPO',
            'CONTROLE',
            'PERGUNTA',
            'ALTERNATIVA',
            'CORRETA',
            'CATEGORIA',
            'SUBCATEGORIA',
            'Dificuldade'
        ]

    @staticmethod
    def verifica_tipo(tipo):
        # Cria um dicionário com todos os tipos da questão
        tipos = {
            'Multipla escolha 1 correta':  'me',
            'Multipla escolha n corretas': 'men',
            'Verdadeiro ou falso':         'vf',
            'Dissertativa':                'd'
        }
        # Retorna a variável que a ser gravada na planilha
        return tipos[tipo]

    def verifica_correta(self, tipo, opc):
        # Dependendo do tipo da questão, temos uma forma diferente de salvar a alternativa correta
        resposta = {
            'vf':  self.vf,
            'men': self.men,
            'me':  self.me,
        }
        return resposta[tipo](opc)

    def me(self, opc):
        # Marca correta caso o indice da opção seja igual ao valor retornado no 'api_dict'
        # Isso porque 'me' retorna um valor numérico, não True e False
        return 'CORRETA' if self.correta == opc else ''

    def vf(self, opc):
        # Marca 'V' caso a opção seja true e 'f' em caso de false
        return 'V' if self.correta[opc] else 'F'

    def men(self, opc):
        # Marca 'CORRETA' caso a opção seja true e '' em caso de false
        return 'CORRETA' if self.correta[opc] else ''

    def salvar_nova_questao(self, linha_a_salvar):
        # Adiciona na última linha disponível do Excel as informações
        self.ws.append(linha_a_salvar)
        # Salva a planilha que até o momento está apenas na memória do (App)

        if not self.diretorio:
            self.diretorio = self.salvar_como()
        self.wb.save(self.diretorio)

    def salvar_edicao(self, linha_a_salvar, opc, linha):
        # Como opção é uma contagem numérica em range, usamos esse valor para acrescentar as linhas, 0 a N, opções
        linha += opc
        coluna = 'ABCDEFGHIJK'
        # Adiciona na última linha disponível do Excel as informações
        for index in range(11):
            self.ws[f'{coluna[index]}{linha}'] = linha_a_salvar[index]
        # Salva a planilha que até o momento está apenas na memória do ‘software’
        self.wb.save(self.diretorio)


# ---------------------------------------------------------------------------------------------------------------------#


class BackEnd(FuncoesBackEnd):
    def gravar(self, api_dict):
        # Inicia planilha do Excel
        self.inicia_excel()

        # Cria um laço para salvar cada alternativa criada
        for opc, alternativa in enumerate(api_dict['Alternativa']):
            # Cria a lista que será usada para salvar na planilha para cada alternativa criada
            linha_a_salvar = self.separa_informacoes(api_dict, opc, alternativa)
            if api_dict['linha']:
                self.salvar_edicao(linha_a_salvar, opc, api_dict['linha'])
            else:
                # Salva a linha na planilha de fato
                self.salvar_nova_questao(linha_a_salvar)

            linha_a_salvar[2] = str(linha_a_salvar[2])
            # Printa no console cada linha que será salva
            print('\t \t'.join(linha_a_salvar))

    def separa_informacoes(self, api_dict, opc, alternativa):
        tipo = self.verifica_tipo(api_dict['Tipo'])
        self.correta = api_dict['Correta']

        # Cria a linha a ser salva com as informações do 'api_dict'
        return [
            '',
            tipo,
            int(api_dict['Peso']),
            api_dict['Tempo'],
            '',
            api_dict['Pergunta'],
            alternativa,
            self.verifica_correta(tipo, opc),
            api_dict['Categoria'],
            api_dict['SubCategoria'],
            api_dict['Dificuldade']
        ]

    def carrega_linha(self, index):
        self.inicia_excel()
        infos = self.ws[f'A{index}:K{index}']
        return tuple(self.get_api_dicio(infos))

    def checa_quantidade_de_opcoes(self, index):
        verifica = self.carrega_linha(index)[5]
        soma = 0
        while True:
            linha = self.carrega_linha(index)[5]
            if verifica != linha:
                break
            index += 1
            soma += 1
        return soma

    @staticmethod
    def get_api_dicio(linha):
        return [linha[0][idx].value for idx in range(len(linha[0]))]

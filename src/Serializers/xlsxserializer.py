from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..Constants import QUESTIOHEADER
from ..Contracts.serializerhandlers import Serializer
from ..Hints.hints import Iterable


class XLSXSerializer(Serializer):
    def export_to_path(self, file_path: Path, data: Iterable) -> None:
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.append([h.upper() for h in QUESTIOHEADER])
        for line_data in data:
            ws.append([line_data[header] for header in QUESTIOHEADER])

        try:
            wb.save(file_path)
        except PermissionError:
            raise PermissionError(
                'A pasta deve estar fechada para salvar uma nova questão.'
            )

    def import_from_path(self, file_path: Path) -> Iterable:
        wb: Workbook = load_workbook(file_path, read_only=True, data_only=True)
        ws: Worksheet = wb.active
        return ws.values

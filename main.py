import os
import os.path
import subprocess

from tkinter import Tk, IntVar, BooleanVar, StringVar, Widget, Text, Button, Checkbutton, Label, Toplevel, Menu, Frame, Entry, Radiobutton
from tkinter.ttk import Combobox
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter.messagebox import showinfo, showwarning, showerror, askyesnocancel

import openpyxl as xlsx

from Modules import DEFAULT_EXTENSION, FILETYPES, SHORTCUTS, LINK_FEEDBACK_FORM
from Modules.configuracoes import Configs
from Modules.place_holder import PlaceHolderEntry
from Modules.question import Question
from Modules.quadro_de_questoes import QuestionFrame
from Modules.corretor_ortografico import PowerfullSpellChecker


BASE_DIR: str = os.path.dirname(__file__)

CONFIGS_DIR: str = os.path.join(BASE_DIR, 'configs')

CONFIGS_FILE: str = os.path.join(CONFIGS_DIR, 'configs.json')

PERFIL_FILE: str = os.path.join(CONFIGS_DIR, 'perfil.json')

SPELLCHECK_FILE: str = os.path.join(CONFIGS_DIR, 'dicionario_pessoal.json')


class Main(Configs, Tk, QuestionFrame):
    __file_path = None
    __option_list = list()
    __exported = True
    __editing = None
    __after_id = None

    def __init__(self):
        Configs.__init__(self, CONFIGS_DIR, CONFIGS_FILE, PERFIL_FILE, SPELLCHECK_FILE)
        self.__init__root()

        self.__var_correct_me_index = IntVar()
        self.__var_index_default_unit = IntVar()
        self.__var_clear_question = BooleanVar()
        self.__var_number_of_questions = StringVar()

        self.__init__frames()
        self.__init__form()
        self.__init__menu()

        QuestionFrame.__init__(self, self.root_configs)

        self.__binds()

        self.mainloop()

    def __init__root(self):
        Tk.__init__(self)
        largura_tela = self.winfo_screenwidth()
        altura_tela = self.winfo_screenheight()
        largura_do_quadro = int(largura_tela * 0.5)
        altura_do_quadro = int(altura_tela * 0.9)
        self.geometry(f'{largura_do_quadro}x{altura_do_quadro}+0+0')
        self.resizable(False, False)
        self.configure(**self.root_configs)
        self.set_titulo()

        self.frame_configs['master'] = self

    def __init__menu(self):
        menubar = Menu(self)
        opcoes_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label='Opções', menu=opcoes_menu, underline=0)

        opcoes_menu.add_command(label='Abrir', command=self.__abrir)
        opcoes_menu.add_command(label='Criar novo', command=self.__salvar_como)

        opcoes_menu.add_separator()

        unidade_padrao = Menu(opcoes_menu, tearoff=0)
        opcoes_menu.add_cascade(label='Unidade Padrão', menu=unidade_padrao)
        for indice, unidade in enumerate(self.unidades):
            unidade_padrao.add_radiobutton(label=unidade, value=indice, variable=self.__var_index_default_unit,
                                           command=self.__altera_unidade_padrao)
            if unidade == self.configuracao_unidade_padrao:
                unidade_padrao.invoke(indice)
        opcoes_menu.add_checkbutton(label='Apagar enunciado ao salvar', variable=self.__var_clear_question,
                                    command=self.__salva_opcao_apagar_enunciado, )
        if self.configuracao_apagar_enunciado:
            opcoes_menu.invoke(4)

        opcoes_menu.add_separator()

        opcoes_menu.add_command(label='Feedbacks', command=self.__abre_feedback)

        help_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label='Ajuda', menu=help_menu, underline=0)

        help_menu.add_command(label='Lista de atalhos', command=self.__abre_atalhos)
        help_menu.add_command(label='Info', command=self.__abrir_infos)

        self.config(menu=menubar)

    def __init__frames(self):
        Frame(name='frame_form', **self.frame_configs).place(relx=0.02, rely=0.02, relwidth=0.96, relheight=0.96)

        self.label_configs['master'] = self.__frame_form
        self.list_configs['master'] = self.__frame_form
        self.buttons_configs['master'] = self.__frame_form
        self.text_configs['master'] = self.__frame_form
        self.entry_configs['master'] = self.__frame_form

    def __init__form(self):
        self.set_quantidade_de_questoes(0)

        Label(**self.label_configs, textvariable=self.__var_number_of_questions).place(relx=0.02, rely=0)

        Label(**self.label_configs, text='Unidade').place(relx=0.2, rely=0)
        Combobox(**self.list_configs, name='unidade', values=self.unidades).place(relx=0.2, rely=0.03, width=165)

        Label(**self.label_configs, text='Código do curso').place(relx=0.5, rely=0)
        Entry(**self.entry_configs, name='codigo_curso').place(relx=0.5, rely=0.03, width=145)

        Label(**self.label_configs, text='Tempo de resposta').place(relx=0.75, rely=0)
        PlaceHolderEntry(**self.entry_configs, name='tempo', placeholder='00:00:00').place(relx=0.75, rely=0.03)

        Label(**self.label_configs, text='Tipo da questão').place(relx=0.2, rely=0.07)
        Combobox(**self.list_configs, name='tipo', values=self.tipos).place(relx=0.2, rely=0.1, width=165)

        Label(**self.label_configs, text='Dificuldade').place(relx=0.5, rely=0.07)
        Combobox(**self.list_configs, name='dificuldade', values=self.dificuldades).place(relx=0.5, rely=0.1)

        Label(**self.label_configs, text='Peso da questão').place(relx=0.75, rely=0.07)
        PlaceHolderEntry(**self.entry_configs, name='peso', placeholder='1').place(relx=0.75, rely=0.1)

        Label(**self.label_configs, text='Enunciado da questão').place(relx=0.05, rely=0.17)
        Text(**self.text_configs, name='pergunta', height=4).place(relx=0.05, rely=0.2, relwidth=0.75)
        self.corretor_pergunta = self.__init_corretor(self.__frame_form.children.get('pergunta'))

        Button(**self.buttons_configs, text='+ opção', name='bt_+opcao',
               command=self.__add_alternativa).place(relx=0.7, rely=0.3, relwidth=0.1)

        Button(**self.buttons_configs, text='- opção', name='bt_-opcao',
               command=self.__remove_alternativa).place(relx=0.85, rely=0.3, relwidth=0.1)

        Button(**self.buttons_configs, text='Exportar', name='bt_exportar',
               command=self.__exportar).place(relx=0.05, rely=0.95, relwidth=0.4)

        Button(**self.buttons_configs, text='Salvar questão', name='bt_salvar',
               command=self.__salvar).place(relx=0.55, rely=0.95, relwidth=0.4)

        self.after(100, self.__frame_form.children['unidade'].focus_set)
        self.__set_opcao('unidade', self.configuracao_unidade_padrao)
        self.__seleciona_opcao('dificuldade', 0)
        self.__seleciona_opcao('tipo', 0)
        self.__organiza_ordem_tabulacao()

    @staticmethod
    def __init_corretor(text_widget: Widget) -> PowerfullSpellChecker:
        return PowerfullSpellChecker(text_widget, timeout=500, SPELLCHECK_FILE=SPELLCHECK_FILE)

    def verifica_digito(self, corretor: PowerfullSpellChecker):
        if self.__after_id:
            self.__frame_form.children['pergunta'].after_cancel(self.__after_id)

        self.__after_id = self.__frame_form.children['pergunta'].after(5000, corretor.__check_spelling)

    def __organiza_ordem_tabulacao(self):
        ordem_widgets = [
            self.__frame_form.children['unidade'],
            self.__frame_form.children['codigo_curso'],
            self.__frame_form.children['tempo'],
            self.__frame_form.children['tipo'],
            self.__frame_form.children['dificuldade'],
            self.__frame_form.children['peso'],
            self.__frame_form.children['pergunta'],
        ]

        if self.__option_list:
            for opcao in self.__option_list:
                ordem_widgets.append(opcao['opcao'])
                ordem_widgets.append(opcao['bt_opcao'][0])

        ordem_widgets.extend([
            self.__frame_form.children['bt_salvar'],
            self.__frame_form.children['bt_exportar'],
            self.__frame_form.children['bt_+opcao'],
            self.__frame_form.children['bt_-opcao'],
        ])

        for widget in ordem_widgets:
            widget.lift()

    def __set_opcao(self, nome_widget: str, opcao: str):
        widget = self.__frame_form.children[nome_widget]
        widget.set(opcao)

    def __seleciona_opcao(self, nome_widget: str, indice: int):
        widget = self.__frame_form.children[nome_widget]
        widget.current(indice)

    def __abrir(self):
        def verifica_correta(certa):
            return certa == 'CORRETA'

        def get_alternativa(work_sheet, line):
            alternativa = work_sheet.cell(row=line, column=7).value
            correta = work_sheet.cell(row=line, column=8).value
            return alternativa, verifica_correta(correta)

        def verifica_tipo(tipo_questao: object) -> object:
            tipos = {
                'me': 'Multipla escolha 1 correta',
                'men': 'Multipla escolha n corretas',
                'vf': 'Verdadeiro ou falso',
                'd': 'Dissertativa',
            }
            return tipos[tipo_questao]

        def get_info(work_sheet, line: object) -> tuple[object, object, object, object, object, object]:
            tipo: object = verifica_tipo(work_sheet.cell(row=line, column=2).value)
            peso: object = work_sheet.cell(row=line, column=3).value
            tempo: object = work_sheet.cell(row=line, column=4).value
            dificuldade: object = work_sheet.cell(row=line, column=11).value
            unidade: object = work_sheet.cell(row=line, column=9).value
            codigo: object = work_sheet.cell(row=line, column=10).value
            return tipo, peso, tempo, dificuldade, unidade, codigo

        if self.__option_list:
            self.__salvar_como()

        self.__exported = False

        self.limpa_tree_view()

        self.__file_path = askopenfilename(defaultextension=DEFAULT_EXTENSION, filetypes=FILETYPES)
        self.set_titulo()

        wb = xlsx.load_workbook(self.__file_path)
        ws = wb.active

        ultima_linha = len(ws['A'])
        controle = None
        opcoes = list()

        for row in range(ultima_linha):
            if not row:
                controle = ws.cell(row=2, column=6).value
                continue
            row += 1
            pergunta = ws.cell(row=row, column=6).value

            if pergunta != controle:
                tipo, peso, tempo, dificuldade, unidade, codigo = get_info(ws, row - 1)
                q = Question(unidade, codigo, tempo, tipo, dificuldade, peso, controle, opcoes)
                self.set_quantidade_de_questoes(self.add_question(q))
                opcoes = list()
                controle = pergunta
                opcoes.append(get_alternativa(ws, row))

            if row == ultima_linha:
                opcoes.append(get_alternativa(ws, row))
                tipo, peso, tempo, dificuldade, unidade, codigo = get_info(ws, row)
                q = Question(unidade, codigo, tempo, tipo, dificuldade, peso, controle, opcoes)
                self.set_quantidade_de_questoes(self.add_question(q))

            if pergunta == controle:
                opcoes.append(get_alternativa(ws, row))

    def __seleciona_caminho(self, titulo: str):
        self.__file_path = asksaveasfilename(confirmoverwrite=True, defaultextension=DEFAULT_EXTENSION, filetypes=FILETYPES, initialdir=BASE_DIR, title=titulo)

    def __salvar_como(self):
        self.__seleciona_caminho('Salvar como')
        self.__exportar()
        self.set_titulo()

    def __abrir_infos(self):
        infos = Toplevel(master=self, **self.root_configs)
        infos.title('Informações')
        infos.geometry('400x300')

        Label(**self.label_configs, text=f'Versão:  {self.versao}').pack(padx=10, pady=10)

        Button(**self.buttons_configs, text='Verificar atualização',
               command=self.__verifica_versao).pack(padx=10, pady=10)

    def __verifica_versao(self):
        pass

    def __add_alternativa(self):
        qnt_opcoes = len(self.__option_list)
        if qnt_opcoes >= 10:
            return showinfo('Limite excedido', 'Limite máximo de opções excedido')

        opcao = Text(**self.text_configs, height=2)
        opcao.place(relx=0.05, rely=self.__calcula_posicao_y(qnt_opcoes), relwidth=0.85)
        self.__option_list.append({'opcao': opcao, 'bt_opcao': self.__constroi_opcao(qnt_opcoes)})
        self.__organiza_ordem_tabulacao()

        corretor_alternativa = self.__init_corretor(opcao)
        opcao.bind('<KeyRelease>', lambda e: corretor_alternativa.start_timer())

    def __remove_alternativa(self):
        if not len(self.__option_list) == 0:
            self.__option_list[-1]['opcao'].destroy()
            self.__option_list[-1]['bt_opcao'][0].destroy()
            self.__option_list.pop()

    def __constroi_opcao(self, valor):
        def cria_radio_bt():
            var = self.__var_correct_me_index
            bt_opcao = Radiobutton(master=self.__frame_form, value=valor, variable=var, **self.bt_opcoes_configs)
            bt_opcao.place(relx=0.9, rely=pos_y)
            return bt_opcao, var

        def cria_check_bt():
            var = BooleanVar()
            bt_opcao = Checkbutton(master=self.__frame_form, anchor='center', **self.bt_opcoes_configs, variable=var)
            bt_opcao.place(relx=0.9, rely=pos_y)
            return bt_opcao, var

        pos_y = self.__calcula_posicao_y(valor) + 0.01

        tipos_btn = {
            'Multipla escolha 1 correta': cria_radio_bt,
            'Multipla escolha n corretas': cria_check_bt,
            'Verdadeiro ou falso': cria_check_bt,
            'Dissertativa': None,
        }
        return tipos_btn[self.__get_tipo]()

    def __altera_alternativa(self):
        if len(self.__option_list):
            for indice, alternativa in enumerate(self.__option_list):
                alternativa['bt_opcao'][0].destroy()
                alternativa.pop('bt_opcao')
                alternativa['bt_opcao'] = self.__constroi_opcao(indice)
        self.__organiza_ordem_tabulacao()

    def __salvar(self):
        self.__exported = False
        if self.__get_pergunta.rstrip('\n') in [questao.question for questao in self.get_lista_questoes]:
            raise showinfo('Pergunta repetida', 'Já existe uma questão com a mesma pergunta')

        if not self.__get_pergunta.rstrip('\n'):
            return showwarning('Pergunta em branco', 'Para salvar uma questão é necessário que a pergunta não esteja '
                                                     'em branco.')

        for opcao in self.__option_list:
            if not opcao['opcao'].get('0.0', 'end').rstrip('\n') or not self.__option_list:
                return showwarning('Opção em branco', 'Para salvar uma questão é necessário que nenhuma opção esteja '
                                                      'em branco e é necesário ter ao menos uma opção.')

        q = Question(self.__get_unidade, self.__get_codigo_curso, self.__get_tempo, self.__get_tipo,
                     self.__get_dificuldade, self.__get_peso, self.__get_pergunta, self.__get_opcoes)

        if self.__editing:
            self.edited_question(q)
        else:
            self.set_quantidade_de_questoes(self.add_question(q))

        self.__limpa_form_parcial()
        self.__editing = None

    def __exportar(self):
        def verifica_correta(correta: bool, tipo: str):
            def verdadeiro_e_falso():
                if correta:
                    return 'V'
                else:
                    return 'F'

            def multipla_escolha():
                if correta:
                    return 'CORRETA'
                else:
                    return ''

            tipos = {
                'Multipla escolha 1 correta': multipla_escolha,
                'Multipla escolha n corretas': multipla_escolha,
                'Verdadeiro ou falso': verdadeiro_e_falso,
            }
            return tipos[tipo]()

        def verifica_tipo(tipo: str):
            tipos = {
                'Multipla escolha 1 correta': 'me',
                'Multipla escolha n corretas': 'men',
                'Verdadeiro ou falso': 'vf',
            }
            return tipos[tipo]

        if not self.__file_path:
            self.__seleciona_caminho('Exportar para')
            print(self.__file_path)

        wb = xlsx.Workbook()
        ws = wb.active
        ws.append([
            'ID', 'TIPO', 'PESO', 'TEMPO', 'CONTROLE', 'PERGUNTA',
            'ALTERNATIVA', 'CORRETA', 'CATEGORIA', 'SUBCATEGORIA', 'Dificuldade'
        ])
        for questao in self.get_lista_questoes:
            for opcao in questao.options:
                ws.append([
                    '', verifica_tipo(questao.type_), questao.weight, questao.time,
                    '', questao.question, opcao[0], verifica_correta(opcao[1], questao.type_),
                    questao.unit, questao.code, questao.difficulty
                ])

        try:
            wb.save(self.__file_path)
        except PermissionError:
            showerror('Não foi possível exportar', 'Para exportar é necessário que o arqivo esteja fechado.'
                                                   '\nConfirme que nenhum arquivo xlsx com mesmo nome esteja aberto.')
        finally:
            self.__exported = True

    def edit_question(self, questao):
        self.__limpa_form_completo()
        self.__editing = questao

        self.__set_opcao('unidade', questao.unit)
        self.__frame_form.children['codigo_curso'].insert(0, questao.code if questao.code else '')
        self.__frame_form.children['tempo'].insert(0, questao.time)
        self.__set_opcao('tipo', questao.type_)
        self.__set_opcao('dificuldade', questao.difficulty)
        self.__frame_form.children['peso'].insert(0, questao.weight)
        self.__frame_form.children['pergunta'].insert(0.0, questao.question)

        for indice, opcao in enumerate(questao.options):
            self.__add_alternativa()
            alternativa = self.__option_list[indice]
            alternativa['opcao'].insert(0.0, opcao[0])
            if opcao[1]:
                alternativa['bt_opcao'][0].select()

    def __binds(self):
        def ctrl_events(e):
            def seleciona_tipo(indice: int):
                self.__seleciona_opcao('tipo', indice - 1)
                self.__altera_alternativa()

            def seleciona_dificuldade(indice: int):
                self.__seleciona_opcao('dificuldade', indice - 4)

            key = e.keysym
            events = {
                'e': self.__exportar,
                's': self.__salvar,
                'o': self.__abrir,
                'equal': self.__add_alternativa,
                'plus': self.__add_alternativa,
                'minus': self.__remove_alternativa,
                '1': seleciona_tipo,
                '2': seleciona_tipo,
                '3': seleciona_tipo,
                '4': seleciona_dificuldade,
                '5': seleciona_dificuldade,
                '6': seleciona_dificuldade,
            }

            if key in events:
                if key.isdigit():
                    return events[key](int(key))
                else:
                    return events[key]()

        def key_events(e):
            key = e.keysym
            events = {
                'F1': self.__abre_atalhos,
                'F4': self.__close_event,
                'F12': self.__salvar_como,
            }
            if key in events.keys():
                return events[key]()

        self.bind('<Control-Key>', lambda e: ctrl_events(e))
        self.bind('<Key>', lambda e: key_events(e))
        self.protocol(self.protocol()[0], self.__close_event)
        self.__frame_form.children['pergunta'].bind('<KeyRelease>', lambda e: self.corretor_pergunta.start_timer())

        self.__frame_form.children['tipo'].bind('<<ComboboxSelected>>', lambda e: self.__altera_alternativa())

    def __close_event(self):
        if not self.__exported:
            tittle = 'Salvamento pendente'
            information = 'Uma ou mais questões estão em edição e não foram exportadas.\n' \
                          'Deseja exportar as alterações antes de sair?'
            answer = askyesnocancel(tittle, information)
            if answer:
                self.__exportar()
                self.quit()
            elif answer is None:
                pass
            else:
                self.quit()
        else:
            self.quit()

    def __limpa_form_parcial(self):
        if self.__var_clear_question.get():
            self.__frame_form.children['pergunta'].delete(0.0, 'end')

        self.__limpa_opcoes()
        self.__frame_form.children['pergunta'].focus_set()

    def __limpa_form_completo(self):
        self.__frame_form.children['tempo'].delete('0', 'end')
        self.__frame_form.children['peso'].delete('0', 'end')
        self.__frame_form.children['pergunta'].delete('0.0', 'end')
        self.__limpa_opcoes()
        self.__frame_form.children['pergunta'].focus_set()

    def __limpa_opcoes(self):
        for opcao in self.__option_list:
            opcao['bt_opcao'][0].destroy()
            opcao['opcao'].destroy()
        self.__option_list.clear()

    def set_titulo(self):
        nome = ''
        if self.__file_path:
            nome = os.path.basename(self.__file_path)
        self.title(f'Gerador de questões - {nome}')

    def __altera_unidade_padrao(self):
        self.__frame_form.children['unidade'].current(self.__var_index_default_unit.get())
        self.salva_informacao('perfil', 'unidade_padrao', self.unidades[self.__var_index_default_unit.get()])

    def __salva_opcao_apagar_enunciado(self):
        self.salva_informacao('perfil', 'apagar_enunciado', self.__var_clear_question.get())

    def set_quantidade_de_questoes(self, param):
        self.__var_number_of_questions.set(f'Número de\nquestões:\n {param}')

    def __abre_atalhos(self):
        top = Toplevel(self, **self.root_configs)
        frame = Frame(top, bg='lightgreen')
        frame.pack(padx=10, pady=10, fill='both')

        formatacao = dict(bg='lightgreen', master=frame)
        posicao = {'padx': 2, 'pady': 5}
        for i, informacoes_atalho in enumerate(SHORTCUTS):
            descricao, atalho = informacoes_atalho
            Label(text=descricao, **formatacao).grid(column=0, row=i, sticky='e', **posicao)
            Label(text=atalho, **formatacao).grid(column=1, row=i, sticky='w', **posicao)

    @staticmethod
    def __calcula_posicao_y(multiplo):
        return 0.35 + (0.06 * multiplo)

    @staticmethod
    def __abre_feedback():
        return subprocess.call(['start', LINK_FEEDBACK_FORM], shell=True, stdout=False)

    @property
    def __frame_form(self):
        return self.children['frame_form']

    @property
    def __get_unidade(self):
        return self.__frame_form.children['unidade'].get()

    @property
    def __get_codigo_curso(self):
        return self.__frame_form.children['codigo_curso'].get()

    @property
    def __get_tempo(self):
        return self.__frame_form.children['tempo'].get()

    @property
    def __get_dificuldade(self):
        return self.__frame_form.children['dificuldade'].get()

    @property
    def __get_peso(self):
        return self.__frame_form.children['peso'].get()

    @property
    def __get_pergunta(self):
        return self.__frame_form.children['pergunta'].get(0.0, 'end-1c')

    @property
    def __get_tipo(self):
        return self.__frame_form.children['tipo'].get()

    @property
    def __get_opcoes(self):
        def verifica_correta(resposta, index):
            if self.__get_tipo == 'Multipla escolha 1 correta':
                return resposta == index
            return resposta

        opcoes = list()
        for indice, dict_opcao in enumerate(self.__option_list):
            alternativa = dict_opcao['opcao'].get(0.0, 'end-1c')
            correta = verifica_correta(dict_opcao['bt_opcao'][1].get(), indice)
            opcoes.append((alternativa, correta))
        return opcoes


if __name__ == '__main__':
    Main()
